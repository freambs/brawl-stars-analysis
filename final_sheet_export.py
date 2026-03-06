import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


# =========================
# CONFIG
# =========================

INPUT_FILE = "map_output.json"
OUTPUT_FILE = "Of RAW DATA.xlsx"


# =========================
# CARREGA PARTIDAS
# =========================

with open(INPUT_FILE, "r", encoding="utf-8") as f:
    data = json.load(f)

all_matches = []

for category in data:
    for mode in data[category]:
        for map_name in data[category][mode]:
            for battle in data[category][mode][map_name]:

                battle_info = battle.get("battle", {})
                teams = battle_info.get("teams", [])
                result = battle_info.get("result", "")
                raw_time = battle.get("battleTime", "")

                if len(teams) != 2 or not raw_time:
                    continue

                if len(teams[0]) != 3 or len(teams[1]) != 3:
                    continue

                try:
                    dt = datetime.strptime(raw_time, "%Y%m%dT%H%M%S.000Z")
                except ValueError:
                    continue

                def extract(team):
                    players = sorted([p.get("tag", "Unknown") for p in team])
                    brawlers = [p.get("brawler", {}).get("name", "Unknown") for p in team]
                    return players, brawlers

                t1_players, t1_brawlers = extract(teams[0])
                t2_players, t2_brawlers = extract(teams[1])

                all_matches.append({
                    "category": category,
                    "mode": mode,
                    "map": map_name,
                    "date": dt,
                    "result": result.lower(),
                    "t1_players": t1_players,
                    "t2_players": t2_players,
                    "t1_brawlers": t1_brawlers,
                    "t2_brawlers": t2_brawlers
                })

print("Total de partidas processadas:", len(all_matches))


# =========================
# ORDENA POR DATA
# =========================

all_matches.sort(key=lambda x: x["date"])


# =========================
# AGRUPA EM MD3
# =========================

sets = []
current_set = []
t1_wins = 0
t2_wins = 0

for match in all_matches:

    if not current_set:
        current_set = [match]
        t1_wins = 1 if match["result"] == "victory" else 0
        t2_wins = 1 if match["result"] != "victory" else 0
        continue

    last = current_set[-1]

    same_map = match["map"] == last["map"]
    same_mode = match["mode"] == last["mode"]
    same_teams = (
        match["t1_players"] == last["t1_players"] and
        match["t2_players"] == last["t2_players"]
    )

    if same_map and same_mode and same_teams:

        current_set.append(match)

        if match["result"] == "victory":
            t1_wins += 1
        else:
            t2_wins += 1

        if t1_wins == 2 or t2_wins == 2:
            sets.append(current_set)
            current_set = []
            t1_wins = 0
            t2_wins = 0

    else:
        sets.append(current_set)
        current_set = [match]
        t1_wins = 1 if match["result"] == "victory" else 0
        t2_wins = 1 if match["result"] != "victory" else 0

if current_set:
    sets.append(current_set)


# =========================
# ABRE OU CRIA PLANILHA
# =========================

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active


# =========================
# EVITA DUPLICATAS
# =========================

existing_rows = set()

for row in ws.iter_rows(values_only=True):
    if row and row[5]:
        signature = str(row[5]) + str(row[2]) + str(row[6])
        existing_rows.add(signature)


# =========================
# EXPORTA DADOS
# =========================

new_rows = 0

for set_group in sets:

    first = set_group[0]

    date_formatted = first["date"].strftime("%d %B %Y") 

    signature = date_formatted + first["map"] + first["t1_players"][0]

    if signature in existing_rows:
        continue

    t1_wins = sum(1 for m in set_group if m["result"] == "victory")
    total_rounds = len(set_group)
    t1_losses = total_rounds - t1_wins

    final_result_t1 = "Victory" if t1_wins > t1_losses else "Defeat"
    final_result_t2 = "Victory" if final_result_t1 == "Defeat" else "Defeat"


    # ======================
    # TIME 1
    # ======================

    row1 = [
        first["mode"],        # A
        "",                   # B (VAZIO)
        first["map"],         # C
        first["category"],    # D
        "",                   # E (VAZIO)
        date_formatted,       # F

        first["t1_players"][0], first["t1_brawlers"][0],
        first["t1_players"][1], first["t1_brawlers"][1],
        first["t1_players"][2], first["t1_brawlers"][2],

        "VS",

        first["t2_players"][0], first["t2_brawlers"][0],
        first["t2_players"][1], first["t2_brawlers"][1],
        first["t2_players"][2], first["t2_brawlers"][2],

        final_result_t1,
        t1_wins,
        t1_losses,
        total_rounds
    ]


    # ======================
    # TIME 2
    # ======================

    row2 = [
        first["mode"],        # A
        "",                   # B (VAZIO)
        first["map"],         # C
        first["category"],    # D
        "",                   # E (VAZIO)
        date_formatted,       # F

        first["t2_players"][0], first["t2_brawlers"][0],
        first["t2_players"][1], first["t2_brawlers"][1],
        first["t2_players"][2], first["t2_brawlers"][2],

        "VS",

        first["t1_players"][0], first["t1_brawlers"][0],
        first["t1_players"][1], first["t1_brawlers"][1],
        first["t1_players"][2], first["t1_brawlers"][2],

        final_result_t2,
        t1_losses,
        t1_wins,
        total_rounds
    ]

    ws.append(row1)
    ws.append(row2)

    new_rows += 2


# =========================
# SALVA PLANILHA
# =========================

wb.save(OUTPUT_FILE)

print("Novas linhas adicionadas:", new_rows)
print("Planilha atualizada com sucesso.")